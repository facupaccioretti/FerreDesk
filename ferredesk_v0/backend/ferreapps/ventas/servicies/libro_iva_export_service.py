"""
Servicio para exportar el Libro IVA Ventas en diferentes formatos.
Genera archivos PDF, Excel y JSON para cumplir con las obligaciones fiscales.
"""

import json
import io
from decimal import Decimal
from typing import Dict, Any, BinaryIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors

# =====================
# Constantes de formato para exportación PDF Libro IVA
# =====================
# Cantidad máxima de comprobantes por página (cada comprobante ocupa 2 filas)
MAX_COMPROBANTES_POR_PAGINA = 14  # Aumentado para aprovechar mejor el espacio
# Altura de página y márgenes (en puntos)
ALTO_PAGINA = A4[1]
MARGEN_SUPERIOR = 0.55 * inch  # Reducido de 1 inch a 0.5 inch
MARGEN_INFERIOR = 0.5 * inch  # Reducido de 1 inch a 0.5 inch
ALTO_TITULO = 16  # Reducido de 20 a 16
ALTO_INFO_PERIODO = 16  # Reducido de 24 a 16
ALTO_ENCABEZADO = 32  # Altura de los encabezados (2 filas)
ALTO_FILA = 18  # Altura de cada fila de comprobante
ALTO_SUBTOTAL = 18  # Altura de la fila de subtotal
ALTO_TRASPASO = 18  # Altura de la fila de traslado

# Definir un gris intermedio para alternar comprobantes
GRIS_COMPROBANTE = colors.Color(0.9, 0.9, 0.9)  # Un gris intermedio, más claro que lightgrey pero más oscuro que whitesmoke


def exportar_libro_iva(formato: str, datos_libro: Dict[str, Any]) -> BinaryIO:
    """
    Exporta el libro IVA en el formato especificado.
    
    Args:
        formato: Formato de exportación ('pdf', 'excel', 'json')
        datos_libro: Datos consolidados del libro IVA
    
    Returns:
        Archivo en memoria con el contenido exportado
    """
    
    if formato.lower() == 'pdf':
        return _generar_pdf_libro_iva(datos_libro)
    elif formato.lower() == 'excel':
        return _generar_excel_libro_iva(datos_libro)
    elif formato.lower() == 'json':
        return _generar_json_libro_iva(datos_libro)
    elif formato.lower() == 'txt':
        return _generar_txt_libro_iva(datos_libro)
    else:
        raise ValueError(f"Formato no soportado: {formato}")


def _generar_pdf_libro_iva(datos_libro: Dict[str, Any]) -> BinaryIO:
    """
    Genera el archivo PDF del libro IVA con formato oficial AFIP, manejando subtotales por página y traspasos.
    """
    import math
    from reportlab.platypus import PageBreak

    # Usar las constantes definidas arriba
    # (No se recalcula MAX_COMPROBANTES_POR_PAGINA aquí, se usa la constante global)

    # Crear buffer en memoria
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=20,
        alignment=1  # Centrado
    )

    # Título y periodo
    periodo = datos_libro['periodo']
    configuracion = datos_libro.get('configuracion', {})
    tipo_libro = configuracion.get('tipo_libro', 'convencional')
    incluir_presupuestos = configuracion.get('incluir_presupuestos', False)
    
    # Determinar título según tipo de libro
    if tipo_libro == 'convencional':
        titulo = f"LIBRO IVA VENTAS (CONVENCIONAL) - {periodo['mes']:02d}/{periodo['anio']}"
    elif tipo_libro == 'administrativo':
        if incluir_presupuestos:
            titulo = f"LIBRO IVA VENTAS (ADMINISTRATIVO CON PRESUPUESTOS) - {periodo['mes']:02d}/{periodo['anio']}"
        else:
            titulo = f"LIBRO IVA VENTAS (ADMINISTRATIVO) - {periodo['mes']:02d}/{periodo['anio']}"
    else:
        titulo = f"LIBRO IVA VENTAS - {periodo['mes']:02d}/{periodo['anio']}"
    
    info_periodo = f"Período: {periodo['mes']:02d}/{periodo['anio']} - Generado: {periodo['fecha_generacion'][:10]}"

    # Encabezados
    headers_datos = ['Fecha', 'Comprobante', 'Número', 'CUIT', 'Razón Social', 'Cond.IVA']
    headers_importes = ['Neto Sin IVA', 'IVA 21%', 'IVA 10.5%', 'IVA 27%', 'Exento', 'Total']
    ancho_pagina = A4[0] - 2*inch
    col_widths = [
        ancho_pagina * 0.13,  # Fecha
        ancho_pagina * 0.13,  # Comprobante
        ancho_pagina * 0.13,  # Número
        ancho_pagina * 0.15,  # CUIT
        ancho_pagina * 0.26,  # Razón Social
        ancho_pagina * 0.20   # Cond.IVA o Total
    ]

    # Preparar filas de comprobantes (cada comprobante son 2 filas)
    filas_comprobantes = []
    for linea in datos_libro['lineas']:
        row_datos = [
            linea['fecha'],
            linea['comprobante'],
            linea['numero'],
            linea['cuit_cliente'],
            linea['razon_social'][:40],
            linea['condicion_iva']
        ]
        row_importes = [
            f"${linea['neto_sin_iva']:,.2f}",
            f"${linea['iva_21']:,.2f}",
            f"${linea['iva_105']:,.2f}",
            f"${linea['iva_27']:,.2f}",
            f"${linea['importe_exento']:,.2f}",
            f"${linea['total']:,.2f}"
        ]
        filas_comprobantes.append((row_datos, row_importes))

    # Dividir en páginas de 4 comprobantes (8 filas)
    total_comprobantes = len(filas_comprobantes)
    paginas = []
    for i in range(0, total_comprobantes, MAX_COMPROBANTES_POR_PAGINA):
        paginas.append(filas_comprobantes[i:i+MAX_COMPROBANTES_POR_PAGINA])

    # Acumuladores de totales
    acumulado_anterior = {
        'neto_sin_iva': Decimal('0.00'),
        'iva_21': Decimal('0.00'),
        'iva_105': Decimal('0.00'),
        'iva_27': Decimal('0.00'),
        'exento': Decimal('0.00'),
        'total': Decimal('0.00')
    }
    # Para el total general
    total_general = acumulado_anterior.copy()

    for idx, pagina in enumerate(paginas):
        # Título y periodo solo en la primera página
        if idx == 0:
            story.append(Paragraph(titulo, title_style))
            story.append(Spacer(1, 10))  # Reducido de 20 a 10
            story.append(Paragraph(info_periodo, styles['Normal']))
            story.append(Spacer(1, 10))  # Reducido de 20 a 10
        else:
            story.append(PageBreak())
            # Al inicio de la página, encabezado de traslado y luego importes
            fila_traspaso_desde = [
                ["Totales trasladados desde página anterior ←", '', '', '', '', ''],
                [
                    f"${acumulado_anterior['neto_sin_iva']:,.2f}",
                    f"${acumulado_anterior['iva_21']:,.2f}",
                    f"${acumulado_anterior['iva_105']:,.2f}",
                    f"${acumulado_anterior['iva_27']:,.2f}",
                    f"${acumulado_anterior['exento']:,.2f}",
                    f"${acumulado_anterior['total']:,.2f}"
                ]
            ]
            tabla_traspaso_desde = Table(fila_traspaso_desde, colWidths=col_widths)
            tabla_traspaso_desde.setStyle(TableStyle([
                ('SPAN', (0, 0), (5, 0)),
                ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (0, 0), 8),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('BACKGROUND', (0, 0), (5, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (5, 0), colors.black),
                # Importes fila
                ('FONTNAME', (0, 1), (5, 1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 1), (5, 1), 8),
                ('ALIGN', (0, 1), (5, 1), 'RIGHT'),
                ('GRID', (0, 0), (5, 1), 0.5, colors.black),
                ('BACKGROUND', (0, 1), (5, 1), colors.whitesmoke),
                ('TEXTCOLOR', (0, 1), (5, 1), colors.black),
            ]))
            story.append(tabla_traspaso_desde)
            story.append(Spacer(1, 10))
        # Tabla de comprobantes de la página
        table_data = [headers_datos, headers_importes]
        subtotal_pagina = {
            'neto_sin_iva': Decimal('0.00'),
            'iva_21': Decimal('0.00'),
            'iva_105': Decimal('0.00'),
            'iva_27': Decimal('0.00'),
            'exento': Decimal('0.00'),
            'total': Decimal('0.00')
        }
        for row_datos, row_importes in pagina:
            table_data.append(row_datos)
            table_data.append(row_importes)
            # Sumar importes a subtotal de página y total general
            subtotal_pagina['neto_sin_iva'] += Decimal(row_importes[0].replace('$','').replace(',',''))
            subtotal_pagina['iva_21'] += Decimal(row_importes[1].replace('$','').replace(',',''))
            subtotal_pagina['iva_105'] += Decimal(row_importes[2].replace('$','').replace(',',''))
            subtotal_pagina['iva_27'] += Decimal(row_importes[3].replace('$','').replace(',',''))
            subtotal_pagina['exento'] += Decimal(row_importes[4].replace('$','').replace(',',''))
            subtotal_pagina['total'] += Decimal(row_importes[5].replace('$','').replace(',',''))
            # Sumar a total general
            total_general['neto_sin_iva'] += Decimal(row_importes[0].replace('$','').replace(',',''))
            total_general['iva_21'] += Decimal(row_importes[1].replace('$','').replace(',',''))
            total_general['iva_105'] += Decimal(row_importes[2].replace('$','').replace(',',''))
            total_general['iva_27'] += Decimal(row_importes[3].replace('$','').replace(',',''))
            total_general['exento'] += Decimal(row_importes[4].replace('$','').replace(',',''))
            total_general['total'] += Decimal(row_importes[5].replace('$','').replace(',',''))
        # Tabla de comprobantes
        table = Table(table_data, repeatRows=2, colWidths=col_widths)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 1), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 1), 8),
            ('ALIGN', (0, 0), (-1, 1), 'CENTER'),
            ('FONTSIZE', (0, 2), (-1, -1), 7),
            ('ALIGN', (0, 2), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            # Líneas verticales en todas las columnas
            ('LINEBEFORE', (0, 0), (0, -1), 0.5, colors.black),
            ('LINEAFTER', (-1, 0), (-1, -1), 0.5, colors.black),
            # Líneas horizontales solo en encabezados y cada dos filas de datos
            ('LINEBELOW', (0, 0), (-1, 1), 0.5, colors.black),  # Encabezados
        ])
        
        # Agregar líneas horizontales cada dos filas de datos (cada comprobante)
        for i in range(2, len(table_data), 2):
            table_style.add('LINEBELOW', (0, i+1), (-1, i+1), 0.5, colors.black)
        for i in range(2, len(table_data), 4):
            if ((i//2) % 2) == 0:  # Ahora el gris va en la segunda, cuarta, etc.
                table_style.add('BACKGROUND', (0, i), (-1, i), GRIS_COMPROBANTE)
                if i+1 < len(table_data):
                    table_style.add('BACKGROUND', (0, i+1), (-1, i+1), GRIS_COMPROBANTE)
        table.setStyle(table_style)
        story.append(table)
        # Solo Totales trasladados, no subtotal
        if idx < len(paginas) - 1:
            fila_traspaso = [
                ["Totales trasladados a la página siguiente →", '', '', '', '', ''],
                [
                    f"${(acumulado_anterior['neto_sin_iva'] + subtotal_pagina['neto_sin_iva']):,.2f}",
                    f"${(acumulado_anterior['iva_21'] + subtotal_pagina['iva_21']):,.2f}",
                    f"${(acumulado_anterior['iva_105'] + subtotal_pagina['iva_105']):,.2f}",
                    f"${(acumulado_anterior['iva_27'] + subtotal_pagina['iva_27']):,.2f}",
                    f"${(acumulado_anterior['exento'] + subtotal_pagina['exento']):,.2f}",
                    f"${(acumulado_anterior['total'] + subtotal_pagina['total']):,.2f}"
                ]
            ]
            tabla_traspaso = Table(fila_traspaso, colWidths=col_widths)
            tabla_traspaso.setStyle(TableStyle([
                ('SPAN', (0, 0), (5, 0)),
                ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (0, 0), 8),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('BACKGROUND', (0, 0), (5, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (5, 0), colors.black),
                # Importes fila
                ('FONTNAME', (0, 1), (5, 1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 1), (5, 1), 8),
                ('ALIGN', (0, 1), (5, 1), 'RIGHT'),
                ('GRID', (0, 0), (5, 1), 0.5, colors.black),
                ('BACKGROUND', (0, 1), (5, 1), colors.whitesmoke),
                ('TEXTCOLOR', (0, 1), (5, 1), colors.black),
            ]))
            story.append(tabla_traspaso)
        # Actualizar acumulado para la próxima página
        for k in acumulado_anterior:
            acumulado_anterior[k] += subtotal_pagina[k]

    # Al final, total general
    story.append(Spacer(1, 10))  # Reducido de 20 a 10
    total_general_data = [
        ['Total Gral.', 'Neto Sin IVA', 'IVA 21%', 'IVA 10.5%', 'IVA 27%', 'Exento'],
        [
            f"${total_general['total']:,.2f}",
            f"${total_general['neto_sin_iva']:,.2f}",
            f"${total_general['iva_21']:,.2f}",
            f"${total_general['iva_105']:,.2f}",
            f"${total_general['iva_27']:,.2f}",
            f"${total_general['exento']:,.2f}"
        ]
    ]
    total_general_table = Table(total_general_data, colWidths=col_widths)
    total_general_style = TableStyle([
        # Encabezados (igual que tabla principal)
        ('FONTNAME', (0, 0), (5, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (5, 0), 8),
        ('ALIGN', (0, 0), (5, 0), 'CENTER'),
        ('BACKGROUND', (0, 0), (5, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (5, 0), colors.black),
        # Filas de datos (igual que tabla principal)
        ('FONTNAME', (0, 1), (5, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (5, 1), 7),
        ('ALIGN', (0, 1), (5, 1), 'RIGHT'),
        ('GRID', (0, 0), (5, 1), 0.5, colors.black),
        ('BACKGROUND', (0, 1), (5, 1), colors.whitesmoke),
        ('TEXTCOLOR', (0, 1), (5, 1), colors.black),
    ])
    total_general_table.setStyle(total_general_style)
    story.append(total_general_table)

    # Generar PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


def _generar_excel_libro_iva(datos_libro: Dict[str, Any]) -> BinaryIO:
    """
    Genera el archivo Excel del libro IVA con fórmulas y formato.
    """
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro IVA Ventas"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    periodo = datos_libro['periodo']
    configuracion = datos_libro.get('configuracion', {})
    tipo_libro = configuracion.get('tipo_libro', 'convencional')
    incluir_presupuestos = configuracion.get('incluir_presupuestos', False)
    
    # Determinar título según tipo de libro
    if tipo_libro == 'convencional':
        titulo_excel = f"LIBRO IVA VENTAS (CONVENCIONAL) - {periodo['mes']:02d}/{periodo['anio']}"
    elif tipo_libro == 'administrativo':
        if incluir_presupuestos:
            titulo_excel = f"LIBRO IVA VENTAS (ADMINISTRATIVO CON PRESUPUESTOS) - {periodo['mes']:02d}/{periodo['anio']}"
        else:
            titulo_excel = f"LIBRO IVA VENTAS (ADMINISTRATIVO) - {periodo['mes']:02d}/{periodo['anio']}"
    else:
        titulo_excel = f"LIBRO IVA VENTAS - {periodo['mes']:02d}/{periodo['anio']}"
    
    ws['A1'] = titulo_excel
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:O1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Información del período
    ws['A3'] = f"Período: {periodo['mes']:02d}/{periodo['anio']}"
    ws['A4'] = f"Generado: {periodo['fecha_generacion'][:10]}"
    
    # Encabezados (fila 6)
    headers = [
        'Fecha', 'Comprobante', 'Número', 'CUIT', 'Nombre/Razón Social',
        'Condición IVA', 'Neto Sin IVA', 'IVA 21%', 'IVA 10.5%', 'IVA 27%',
        'IVA Otras', 'Exento', 'Total'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Datos
    row_num = 7
    for linea in datos_libro['lineas']:
        ws.cell(row=row_num, column=1, value=linea['fecha'])
        ws.cell(row=row_num, column=2, value=linea['comprobante'])
        ws.cell(row=row_num, column=3, value=linea['numero'])
        ws.cell(row=row_num, column=4, value=linea['cuit_cliente'])
        ws.cell(row=row_num, column=5, value=linea['razon_social'])
        ws.cell(row=row_num, column=6, value=linea['condicion_iva'])
        ws.cell(row=row_num, column=7, value=float(linea['neto_sin_iva']))
        ws.cell(row=row_num, column=8, value=float(linea['iva_21']))
        ws.cell(row=row_num, column=9, value=float(linea['iva_105']))
        ws.cell(row=row_num, column=10, value=float(linea['iva_27']))
        ws.cell(row=row_num, column=11, value=float(linea['iva_otras']))
        ws.cell(row=row_num, column=12, value=float(linea['importe_exento']))
        ws.cell(row=row_num, column=13, value=float(linea['total']))
        
        # Aplicar bordes a toda la fila
        for col in range(1, 14):
            ws.cell(row=row_num, column=col).border = border
        
        row_num += 1
    
    # Fórmulas de totales (fila después de los datos)
    total_row = row_num + 1
    ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
    
    # Fórmulas SUM para cada columna numérica
    for col in [7, 8, 9, 10, 11, 12, 13]:
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}7:{col_letter}{row_num})"
        ws.cell(row=total_row, column=col, value=formula)
        ws.cell(row=total_row, column=col).font = Font(bold=True)
        ws.cell(row=total_row, column=col).border = border
    
    # Ajustar ancho de columnas
    column_widths = [12, 12, 10, 15, 30, 15, 12, 12, 12, 12, 12, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Formato de moneda para columnas numéricas
    for row in range(7, row_num + 1):
        for col in [7, 8, 9, 10, 11, 12, 13]:
            cell = ws.cell(row=row, column=col)
            cell.number_format = '#,##0.00'
    
    # Crear buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _generar_json_libro_iva(datos_libro: Dict[str, Any]) -> BinaryIO:
    """
    Genera el archivo JSON del libro IVA para integración con otros sistemas.
    """
    # Convertir Decimal a float para serialización JSON
    def decimal_to_float(obj):
        if isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, dict):
            return {key: decimal_to_float(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [decimal_to_float(item) for item in obj]
        return obj
    # Convertir datos
    datos_serializables = decimal_to_float(datos_libro)
    # Crear buffer
    buffer = io.BytesIO()
    json_str = json.dumps(datos_serializables, indent=2, ensure_ascii=False)
    buffer.write(json_str.encode('utf-8'))
    buffer.seek(0)
    return buffer


# =====================
# Exportación TXT (ARCA/AFIP)
# =====================

# Constantes configurables para TXT
# Código de moneda por defecto para pesos argentinos
MONEDA_POR_DEFECTO = 'PES'
# Cotización para pesos (6 decimales)
COTIZACION_PESOS = '1.000000'
# Código de operación habitual
CODIGO_OPERACION_DEF = '0'
# Codificación del archivo TXT. ARCA/portales suelen requerir ANSI (Windows-1252)
CODIFICACION_TXT = 'windows-1252'


def _generar_txt_libro_iva(datos_libro: Dict[str, Any]) -> BinaryIO:
    """
    Genera el archivo TXT del libro IVA Ventas con separador '|',
    cumpliendo el layout típico AFIP/ARCA (comprobantes + columnas principales).
    """
    buffer = io.BytesIO()

    def formatear_fecha_yyyymmdd(fecha_iso: str) -> str:
        # Espera 'YYYY-MM-DD' y retorna 'YYYYMMDD'
        if not fecha_iso or len(fecha_iso) < 10:
            return ''
        try:
            return fecha_iso.replace('-', '')[:8]
        except Exception:
            return ''

    def pad_izquierda(valor: int | str | None, ancho: int) -> str:
        try:
            n = int(valor) if valor is not None else 0
        except Exception:
            n = 0
        return str(n).zfill(ancho)

    def limpiar_numeros(cadena: str | None) -> str:
        if not cadena:
            return ''
        return ''.join(ch for ch in str(cadena) if ch.isdigit())

    def determinar_doc_tipo_y_nro(cuit: str | None) -> tuple[str, str]:
        cuit_limpio = limpiar_numeros(cuit)
        if len(cuit_limpio) == 11:
            return '80', cuit_limpio  # CUIT
        # Si no hay CUIT, consumidor final sin identificación
        return '99', '0'

    def dec2(valor) -> str:
        # Convierte Decimal/float/str a string con 2 decimales y punto
        if valor is None:
            valor = Decimal('0.00')
        if isinstance(valor, (int, float)):
            valor = Decimal(str(valor))
        if isinstance(valor, str):
            try:
                valor = Decimal(valor)
            except Exception:
                valor = Decimal('0.00')
        return format(valor.quantize(Decimal('0.01')), 'f')

    # Detectar si es Nota de Crédito por código AFIP
    def es_nota_credito(codigo_afip: str | None) -> bool:
        return str(codigo_afip) in {'003', '008', '012'}

    lineas_salida: list[str] = []

    for linea in datos_libro.get('lineas', []):
        # Campos básicos
        tipo_cbte = str(linea.get('comprobante_codigo_afip') or '')
        pto_vta = pad_izquierda(linea.get('ven_punto'), 4)
        nro = pad_izquierda(linea.get('ven_numero'), 8)
        fecha = formatear_fecha_yyyymmdd(linea.get('fecha'))

        doc_tipo, doc_nro = determinar_doc_tipo_y_nro(linea.get('cuit_cliente'))

        razon_social = (linea.get('razon_social') or '').strip()
        moneda = MONEDA_POR_DEFECTO
        cotiz = COTIZACION_PESOS
        codigo_operacion = CODIGO_OPERACION_DEF

        # Importes base tomados de la línea (ya consolidados)
        neto_21 = linea.get('neto_21') or Decimal('0.00')
        iva_21 = linea.get('iva_21') or Decimal('0.00')
        neto_105 = linea.get('neto_105') or Decimal('0.00')
        iva_105 = linea.get('iva_105') or Decimal('0.00')
        neto_27 = linea.get('neto_27') or Decimal('0.00')
        iva_27 = linea.get('iva_27') or Decimal('0.00')
        exento = linea.get('importe_exento') or Decimal('0.00')
        no_gravado = linea.get('no_gravado') or Decimal('0.00')
        total = linea.get('total') or Decimal('0.00')

        # Reglas por letra y tipo de comprobante
        letra = (linea.get('comprobante_letra') or '').strip().upper()
        codigo = tipo_cbte

        # Factura C: total íntegro en No Gravado, netos/IVA a 0
        if letra == 'C':
            no_gravado = total
            neto_21 = Decimal('0.00')
            iva_21 = Decimal('0.00')
            neto_105 = Decimal('0.00')
            iva_105 = Decimal('0.00')
            neto_27 = Decimal('0.00')
            iva_27 = Decimal('0.00')
            exento = Decimal('0.00')

        # Notas de crédito: importes en negativo
        if es_nota_credito(codigo):
            neto_21 = -abs(Decimal(neto_21))
            iva_21 = -abs(Decimal(iva_21))
            neto_105 = -abs(Decimal(neto_105))
            iva_105 = -abs(Decimal(iva_105))
            neto_27 = -abs(Decimal(neto_27))
            iva_27 = -abs(Decimal(iva_27))
            exento = -abs(Decimal(exento))
            no_gravado = -abs(Decimal(no_gravado))
            total = -abs(Decimal(total))

        # Percepciones e impuestos internos (0 por defecto si no se tienen fuentes)
        percepcion_iva = Decimal('0.00')
        percepcion_iibb = Decimal('0.00')
        impuestos_internos = Decimal('0.00')

        cae = linea.get('ven_cae') or ''
        f_vto_cae = linea.get('ven_caevencimiento')
        if f_vto_cae:
            try:
                f_vto_cae = str(f_vto_cae)
                fecha_vto_cae = f_vto_cae.replace('-', '')[:8]
            except Exception:
                fecha_vto_cae = ''
        else:
            fecha_vto_cae = ''

        campos = [
            tipo_cbte,
            pto_vta,
            nro,
            fecha,
            doc_tipo,
            doc_nro,
            razon_social,
            moneda,
            cotiz,
            codigo_operacion,
            dec2(neto_21),
            dec2(iva_21),
            dec2(neto_105),
            dec2(iva_105),
            dec2(neto_27),
            dec2(iva_27),
            dec2(exento),
            dec2(no_gravado),
            dec2(percepcion_iva),
            dec2(percepcion_iibb),
            dec2(impuestos_internos),
            dec2(total),
            str(cae),
            fecha_vto_cae,
        ]

        linea_txt = '|'.join(str(c) for c in campos)
        lineas_salida.append(linea_txt)

    contenido = ('\r\n'.join(lineas_salida) + '\r\n') if lineas_salida else ''
    buffer.write(contenido.encode(CODIFICACION_TXT, errors='replace'))
    buffer.seek(0)
    return buffer

def obtener_nombre_archivo(formato: str, mes: int, anio: int, tipo_libro: str = 'convencional', incluir_presupuestos: bool = False) -> str:
    """
    Genera el nombre del archivo de exportación.
    
    Args:
        formato: Formato del archivo
        mes: Mes del período
        anio: Año del período
        tipo_libro: Tipo de libro ('convencional' o 'informal')
        incluir_presupuestos: Si incluye presupuestos (solo para informal)
    
    Returns:
        Nombre del archivo
    """
    
    # Determinar sufijo según tipo de libro
    if tipo_libro == 'convencional':
        sufijo = 'Convencional'
    elif tipo_libro == 'administrativo':
        if incluir_presupuestos:
            sufijo = 'AdministrativoConPresupuestos'
        else:
            sufijo = 'Administrativo'
    else:
        sufijo = ''
    
    nombre_base = f"Libro_IVA_Ventas_{sufijo}_{mes:02d}{anio}"
    
    if formato.lower() == 'pdf':
        return f"{nombre_base}.pdf"
    elif formato.lower() == 'excel':
        return f"{nombre_base}.xlsx"
    elif formato.lower() == 'json':
        return f"{nombre_base}.json"
    elif formato.lower() == 'txt':
        return f"{nombre_base}.txt"
    else:
        return f"{nombre_base}.{formato}" 